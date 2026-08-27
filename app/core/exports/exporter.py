"""
Export layer: streams job results out to CSV / JSON / JSONL / XLSX.

Takes an iterator of result dicts (see Database.iter_all_results) rather
than a materialized list, so exporting a 200k-row job doesn't require
holding it all in RAM at once (spec section 29, "performance").
"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Callable, Iterable, Iterator, Optional

ProgressCB = Optional[Callable[[int], None]]  # called with rows-written-so-far


def _rows(results: Iterable[dict]) -> Iterator[dict]:
    """Each stored result row is {id, job_id, source_url, data_json, scraped_at}.
    Flatten it to {**data, source_url, scraped_at} for export."""
    for r in results:
        data = json.loads(r["data_json"]) if isinstance(r.get("data_json"), str) else dict(r.get("data", {}))
        flat = dict(data)
        flat["source_url"] = r.get("source_url")
        flat["scraped_at"] = r.get("scraped_at")
        yield flat


def export_csv(results: Iterable[dict], dest: str | Path, on_progress: ProgressCB = None) -> int:
    dest = Path(dest)
    rows = list(_rows(results))
    fieldnames = _collect_fieldnames(rows)
    count = 0
    with dest.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: _stringify(v) for k, v in row.items()})
            count += 1
            if on_progress and count % 100 == 0:
                on_progress(count)
    if on_progress:
        on_progress(count)
    return count


def export_json(results: Iterable[dict], dest: str | Path, on_progress: ProgressCB = None) -> int:
    dest = Path(dest)
    rows = list(_rows(results))
    with dest.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    if on_progress:
        on_progress(len(rows))
    return len(rows)


def export_jsonl(results: Iterable[dict], dest: str | Path, on_progress: ProgressCB = None) -> int:
    dest = Path(dest)
    count = 0
    with dest.open("w", encoding="utf-8") as f:
        for row in _rows(results):
            f.write(json.dumps(row, ensure_ascii=False))
            f.write("\n")
            count += 1
            if on_progress and count % 100 == 0:
                on_progress(count)
    if on_progress:
        on_progress(count)
    return count


def export_xlsx(results: Iterable[dict], dest: str | Path, on_progress: ProgressCB = None) -> int:
    from openpyxl import Workbook

    rows = list(_rows(results))
    fieldnames = _collect_fieldnames(rows)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Results")
    ws.append(fieldnames)
    count = 0
    for row in rows:
        ws.append([_stringify(row.get(k)) for k in fieldnames])
        count += 1
        if on_progress and count % 100 == 0:
            on_progress(count)
    wb.save(str(dest))
    if on_progress:
        on_progress(count)
    return count


# ---------------------------------------------------------------------------
# Odoo CRM Lead import format - matches the exact column layout of Odoo's own
# "crm.lead" import template (the file a user downloads from Settings ->
# Import Records -> "crm_lead_1.xls" inside Odoo, or CRM -> Leads -> Import ->
# "Download Import Template"), including its "Import FAQ" sheet, so a file
# LOGY exports can be dropped straight into that same Odoo import screen with
# no manual re-mapping. See ODOO_COLUMNS below for the exact header order.
# ---------------------------------------------------------------------------
ODOO_COLUMNS = [
    "External ID", "Name", "Company Name", "Contact Name", "Email",
    "Job Position", "Phone", "Mobile", "Street", "Street2", "City",
    "State", "Zip", "Country", "Website", "Notes", "Channel",
]

# "Channel" isn't part of Odoo's own stock crm.lead import template -
# it's a REQUIRED field only on THIS user's own Odoo instance (a custom
# field their install added, not something LOGY can know the valid
# values for by guessing - "مينفعش أخمن قيمة غلط هتبوظ الامبورت"). Every
# exported row gets the SAME fixed value from Settings -> "Odoo Export" ->
# "Channel value" (see SettingsScreen / db setting "odoo_channel_value"),
# defaulting to "Website" per the user's own answer, unless the scraped
# data already has a field literally called "channel" (checked first, in
# _lead_row_for_odoo() below, same as every other Odoo column).
DEFAULT_ODOO_CHANNEL_VALUE = "Website"

# One or more scraped-field names (case-insensitive) that map to each Odoo
# column. Scraped field names come from whatever the user's own Field
# Builder / AI Auto-Extract called them, so this is a best-effort guess
# rather than a fixed schema - first matching key wins per column.
_ODOO_FIELD_MAP: dict[str, tuple[str, ...]] = {
    "Name": ("name", "title", "lead_name", "opportunity"),
    "Company Name": ("company_name", "company", "business_name", "business"),
    "Contact Name": ("contact_name", "full_name", "owner_name", "contact"),
    "Email": ("email", "e-mail", "mail", "owner_email", "contact_email"),
    "Job Position": ("job_position", "position", "job_title", "title"),
    "Phone": ("phone", "phone_number", "telephone", "tel"),
    "Mobile": ("mobile", "owner_phone", "cell"),
    "Street": ("street", "address", "address1"),
    "Street2": ("street2", "address2"),
    "City": ("city",),
    "State": ("state", "region", "province"),
    "Zip": ("zip", "zipcode", "postal_code", "postcode"),
    "Country": ("country", "country_code"),
    "Website": ("website", "site", "url", "web", "homepage"),
    "Notes": ("notes", "description", "note", "comment", "comments"),
    "Channel": ("channel", "lead_channel", "source_channel"),
}


def _lead_row_for_odoo(flat: dict, external_id: str, channel_value: str = DEFAULT_ODOO_CHANNEL_VALUE) -> list:
    lower = {str(k).lower(): v for k, v in flat.items()}
    row = []
    for col in ODOO_COLUMNS:
        if col == "External ID":
            row.append(external_id)
            continue
        value = None
        for key in _ODOO_FIELD_MAP.get(col, ()):
            if key in lower and lower[key] not in (None, ""):
                value = lower[key]
                break
        if col == "Channel" and value is None:
            value = channel_value  # the lead itself has no "channel" field - fall back to the fixed setting
        row.append(_stringify(value) if value is not None else "")
    return row


# The "Import FAQ" sheet reproduced verbatim from Odoo's own crm.lead import
# template, so the exported file guides a user through Odoo's import screen
# exactly the way the original template does.
_ODOO_IMPORT_FAQ = [
    "How to customize the file?",
    "Add, remove and sort columns as you want.",
    "Keep the header (first row) as is for columns you need to keep. Those column labels will be automatically matched in Odoo.",
    "Put any title to your new columns. You can select the fields to match when importing in Odoo.",
    "Mandatory fields to import are the mandatory fields not populated with default values through the system.",
    "It is not recommended to remove the 'ID column' (see here below).",
    "",
    "How to import this file?",
    "Keep the file type as '.xls' as fields formatting is automatic.",
    "If you import in '.csv', double check that the formatting is correctly interpreted in Odoo (encoding format, date format, separators, etc.).",
    "",
    "What is the 'External ID' for?",
    "External ID are unique identifiers for imported records.",
    "If an ID is set to every record, you can reimport the same file several time and Odoo will update records instead of creating new ones if the ID already exists.",
    "You can create your own ID sequence or use the one of your previous software to ease the migration process. Otherwise extend the structure suggested in this file to all your records (by simple drag).",
    "Using ID,  you can safely point out related records from other database tables (e.g. vendors or tags when importing products). You can also use the record name but the import will stop in case of several matching.",
    "",
    "How to import many2one and many2many relationships (e.g. tags)?",
    "Use value names or IDs.",
    'You can create such related records on the fly if they don\'t exist by checking the box "Create if doesn\'t exist" showing up in the column.',
    "To import several m2m values, separate them with a comma without any spacing (e.g. for customer tags: B2B,Medium Size,Clothes Industry).",
    "",
    "How to import one2many relationships (e.g. orders or invoices with several lines)?",
    'To import fields of one2many fields (unit price, quantity, etc. of order/invoice/pricelist lines), make sure "Show fields of relation fields" is checked in the import interface. It allows to map unmatched one2many columns.',
    "You need to reserve one row for each one2many record. Fields of parent level must be left empty so that the system knows there are several o2m lines to import for the same record.",
    "",
    "How to import translated values for translatable fields?",
    "If you install several languages in Odoo, you can set translations for your master data (product names, descriptions, etc.) by clicking the little Earth icon showing up in the field and in the mapping zone of the import screen.",
    "In the import screen, you will be suggested to choose the translations to target with the column to import.",
    "You can therefore use several columns for the same field, each of them pointing specific translations.",
    "",
    "Need more information? Check out our full FAQ at:",
    "https://www.odoo.com/documentation/user/online/general/base_import/import_faq.html",
]


def export_odoo_xlsx(
    results: Iterable[dict], dest: str | Path, on_progress: ProgressCB = None,
    channel_value: str = DEFAULT_ODOO_CHANNEL_VALUE,
) -> int:
    """Exports in the same format/layout as Odoo's own CRM Lead import
    template (External ID / Name / Company Name / Contact Name / Email /
    Job Position / Phone / Mobile / Street / Street2 / City / State / Zip /
    Country / Website / Notes / Channel, plus an 'Import FAQ' sheet) so the
    output can be imported straight into Odoo's CRM -> Leads -> Import
    screen. `channel_value` fills the "Channel" column for every row that
    has no scraped "channel" field of its own - see DEFAULT_ODOO_CHANNEL_VALUE
    above and Settings -> "Odoo Export" for where a user changes it without
    editing this file."""
    from openpyxl import Workbook

    dest = Path(dest)
    rows = list(_rows(results))

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(ODOO_COLUMNS)
    count = 0
    for i, row in enumerate(rows, start=1):
        ws.append(_lead_row_for_odoo(row, f"crm_lead_{i}", channel_value))
        count += 1
        if on_progress and count % 100 == 0:
            on_progress(count)

    faq = wb.create_sheet("Import FAQ")
    for line in _ODOO_IMPORT_FAQ:
        faq.append([line])

    wb.save(str(dest))
    if on_progress:
        on_progress(count)
    return count


def export_odoo_xls(
    results: Iterable[dict], dest: str | Path, on_progress: ProgressCB = None,
    channel_value: str = DEFAULT_ODOO_CHANNEL_VALUE,
) -> int:
    """Same layout/columns as export_odoo_xlsx() above, but written as a
    REAL legacy .xls (BIFF8 binary) file instead of .xlsx (OOXML) -
    "عايزه يطلع XLS مش XSLS": Odoo's own downloadable CRM Lead import
    template is itself a "crm_lead 1.xls" file (see this file's module
    docstring / the very first file the user uploaded), and some Odoo
    versions/import flows are pickier about accepting that exact legacy
    format over .xlsx. openpyxl (used above) can only WRITE .xlsx - it
    dropped .xls support entirely - so this uses xlwt instead, the
    library that still writes genuine legacy Excel binary files. This is
    NOT just export_odoo_xlsx()'s file renamed to '.xls' (that would be
    an .xlsx file wearing the wrong extension, which real Excel/Odoo can
    still detect and may reject or mis-parse) - the bytes on disk are an
    actual .xls workbook.

    xlwt is an old, no-longer-updated library (last released 2019) but
    is still the correct tool for this: the legacy .xls format itself
    hasn't changed, and no actively-maintained library replaced it for
    WRITING (only reading, e.g. xlrd). Its one real limitation versus
    openpyxl is a 65,536-row-per-sheet ceiling (the old Excel format's
    own limit, not something LOGY imposes) - fine for the Odoo import use
    case (Odoo's own bulk-import guidance recommends batches well under
    that anyway), but worth knowing if this function is ever reused
    elsewhere for very large exports."""
    try:
        import xlwt
    except ImportError as e:
        raise ImportError(
            "تصدير Odoo بصيغة .xls محتاج مكتبة xlwt ومش متثبتة. افتح الطرفية وشغّل:\n\n"
            "pip install xlwt\n\nوبعدين جرّب التصدير تاني."
        ) from e

    dest = Path(dest)
    rows = list(_rows(results))

    wb = xlwt.Workbook()
    ws = wb.add_sheet("Template")
    for col_i, col_name in enumerate(ODOO_COLUMNS):
        ws.write(0, col_i, col_name)
    count = 0
    for i, row in enumerate(rows, start=1):
        values = _lead_row_for_odoo(row, f"crm_lead_{i}", channel_value)
        for col_i, value in enumerate(values):
            ws.write(i, col_i, value)
        count += 1
        if on_progress and count % 100 == 0:
            on_progress(count)

    faq = wb.add_sheet("Import FAQ")
    for row_i, line in enumerate(_ODOO_IMPORT_FAQ):
        faq.write(row_i, 0, line)

    wb.save(str(dest))
    if on_progress:
        on_progress(count)
    return count


EXPORTERS = {
    "csv": export_csv,
    "json": export_json,
    "jsonl": export_jsonl,
    "xlsx": export_xlsx,
    "odoo_xlsx": export_odoo_xlsx,
    "odoo_xls": export_odoo_xls,
}


def export(fmt: str, results: Iterable[dict], dest: str | Path, on_progress: ProgressCB = None, **kwargs) -> int:
    fmt = fmt.lower().lstrip(".")
    if fmt not in EXPORTERS:
        raise ValueError(f"صيغة تصدير غير مدعومة: {fmt}. المتاح: {', '.join(EXPORTERS)}")
    # **kwargs lets a caller pass exporter-specific options (currently
    # only export_odoo_xlsx's channel_value - see new_scrape.py's
    # _export_results()) without every OTHER exporter needing to accept
    # and ignore them.
    return EXPORTERS[fmt](results, dest, on_progress, **kwargs)


def _collect_fieldnames(rows: list[dict]) -> list[str]:
    names: list[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                names.append(key)
    return names


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)
