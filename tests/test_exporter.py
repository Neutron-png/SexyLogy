import csv
import json
import tempfile
from pathlib import Path

from app.core.exports import exporter

SAMPLE_ROWS = [
    {"job_id": 1, "source_url": "https://x.com/1", "data_json": json.dumps({"name": "A", "price": "10"}), "scraped_at": 1.0},
    {"job_id": 1, "source_url": "https://x.com/2", "data_json": json.dumps({"name": "B", "price": "20"}), "scraped_at": 2.0},
]


def test_export_csv():
    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "out.csv"
        count = exporter.export_csv(SAMPLE_ROWS, dest)
        assert count == 2
        with dest.open(encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["name"] == "A"
        assert rows[1]["price"] == "20"
        assert "source_url" in rows[0]


def test_export_json():
    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "out.json"
        count = exporter.export_json(SAMPLE_ROWS, dest)
        assert count == 2
        data = json.loads(dest.read_text(encoding="utf-8"))
        assert data[0]["name"] == "A"


def test_export_jsonl():
    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "out.jsonl"
        count = exporter.export_jsonl(SAMPLE_ROWS, dest)
        assert count == 2
        lines = dest.read_text(encoding="utf-8").strip().splitlines()
        assert len(lines) == 2
        assert json.loads(lines[0])["name"] == "A"


def test_export_xlsx():
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "out.xlsx"
        count = exporter.export_xlsx(SAMPLE_ROWS, dest)
        assert count == 2
        wb = load_workbook(dest)
        ws = wb["Results"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0][0] == "name"  # header
        assert "A" in rows[1]


ODOO_SAMPLE_ROWS = [
    {
        "job_id": 1, "source_url": "https://x.com/1", "scraped_at": 1.0,
        "data_json": json.dumps({
            "name": "Need to customize the solution",
            "company_name": "E-light Industry",
            "contact_name": "Henry Jordan",
            "email": "henry@elight.com",
            "phone": "+1 650 356 451",
            "city": "Buenos Aires",
            "country": "AR",
        }),
    },
    {
        "job_id": 1, "source_url": "https://x.com/2", "scraped_at": 2.0,
        "data_json": json.dumps({"name": "Resource Planning project development", "city": "Birmingham"}),
    },
]


def test_export_odoo_xlsx_matches_template_layout():
    from openpyxl import load_workbook
    from app.core.exports.exporter import ODOO_COLUMNS

    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "crm_leads.xlsx"
        count = exporter.export_odoo_xlsx(ODOO_SAMPLE_ROWS, dest)
        assert count == 2

        wb = load_workbook(dest)
        assert wb.sheetnames == ["Template", "Import FAQ"]

        ws = wb["Template"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == ODOO_COLUMNS

        row1 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        assert row1[0] == "crm_lead_1"  # External ID
        assert row1[2] == "E-light Industry"  # Company Name
        assert row1[4] == "henry@elight.com"  # Email

        row2 = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
        assert row2[0] == "crm_lead_2"

        faq = wb["Import FAQ"]
        faq_lines = [r[0].value for r in faq.iter_rows()]
        assert faq_lines[0] == "How to customize the file?"
        assert any("External ID" in (line or "") for line in faq_lines)


def test_export_dispatch_routes_odoo_xlsx():
    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "out.xlsx"
        count = exporter.export("odoo_xlsx", ODOO_SAMPLE_ROWS, dest)
        assert count == 2


def test_export_odoo_xlsx_channel_defaults_and_is_overridable():
    """'Missing required value for the field Channel' - some Odoo
    instances have a CUSTOM required 'Channel' field that isn't part of
    the stock crm.lead import template at all, so LOGY can't know its
    valid values by guessing (see exporter.py's DEFAULT_ODOO_CHANNEL_VALUE
    docstring) - it's a single fixed value applied to every exported row,
    settable from Settings -> 'Odoo Export' without editing code."""
    from openpyxl import load_workbook
    from app.core.exports.exporter import DEFAULT_ODOO_CHANNEL_VALUE

    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "default.xlsx"
        exporter.export_odoo_xlsx(ODOO_SAMPLE_ROWS, dest)
        ws = load_workbook(dest)["Template"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        channel_col = header.index("Channel")
        row1 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        assert row1[channel_col] == DEFAULT_ODOO_CHANNEL_VALUE

        dest2 = Path(tmp) / "custom.xlsx"
        exporter.export_odoo_xlsx(ODOO_SAMPLE_ROWS, dest2, channel_value="Cold Call")
        ws2 = load_workbook(dest2)["Template"]
        row1b = [c.value for c in next(ws2.iter_rows(min_row=2, max_row=2))]
        assert row1b[channel_col] == "Cold Call"

        # A lead that already has its OWN "channel" field wins over the
        # fixed setting - same precedence as every other Odoo column.
        dest3 = Path(tmp) / "own_channel.xlsx"
        rows_with_channel = [{
            "job_id": 1, "source_url": "https://x.com/1", "scraped_at": 1.0,
            "data_json": json.dumps({"name": "Lead", "channel": "Referral"}),
        }]
        exporter.export_odoo_xlsx(rows_with_channel, dest3, channel_value="Website")
        ws3 = load_workbook(dest3)["Template"]
        row1c = [c.value for c in next(ws3.iter_rows(min_row=2, max_row=2))]
        assert row1c[channel_col] == "Referral"


def test_export_odoo_xls_writes_real_legacy_binary_format():
    """'عايزه يطلع XLS مش XSLS' - export_odoo_xls() must write an actual
    legacy .xls (BIFF8) file via xlwt, not just export_odoo_xlsx()'s
    .xlsx bytes saved under a '.xls' name (real Excel/Odoo can tell the
    difference - see export_odoo_xls()'s docstring). xlwt isn't
    installable in this sandbox (no network access to PyPI - same
    constraint noted in tests/run_tests.py's own module docstring), so
    this skips itself if it's genuinely missing rather than failing the
    whole suite; it exercises the real writer end-to-end wherever xlwt
    IS available (including the user's own machine, where `pip install
    xlwt` per requirements.txt works normally)."""
    try:
        import xlwt  # noqa: F401
        import xlrd  # only used here, to read the .xls back for assertions
    except ImportError:
        return  # xlwt/xlrd unavailable in this environment - see docstring above

    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "crm_leads.xls"
        count = exporter.export_odoo_xls(ODOO_SAMPLE_ROWS, dest, channel_value="Website")
        assert count == 2

        book = xlrd.open_workbook(str(dest))
        assert book.sheet_names() == ["Template", "Import FAQ"]

        sheet = book.sheet_by_name("Template")
        header = sheet.row_values(0)
        from app.core.exports.exporter import ODOO_COLUMNS
        assert header == ODOO_COLUMNS

        row1 = sheet.row_values(1)
        assert row1[0] == "crm_lead_1"  # External ID
        assert row1[2] == "E-light Industry"  # Company Name
        assert row1[ODOO_COLUMNS.index("Channel")] == "Website"

        faq = book.sheet_by_name("Import FAQ")
        assert faq.cell_value(0, 0) == "How to customize the file?"


def test_export_dispatch_passes_channel_value_kwarg_to_odoo_exporter():
    with tempfile.TemporaryDirectory() as tmp:
        dest = Path(tmp) / "out.xlsx"
        from openpyxl import load_workbook
        exporter.export("odoo_xlsx", ODOO_SAMPLE_ROWS, dest, channel_value="Direct")
        ws = load_workbook(dest)["Template"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        channel_col = header.index("Channel")
        row1 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        assert row1[channel_col] == "Direct"


def test_export_dispatch_unknown_format_raises():
    try:
        exporter.export("pdf", SAMPLE_ROWS, "out.pdf")
        assert False
    except ValueError:
        pass


def test_export_progress_callback_called():
    calls = []
    with tempfile.TemporaryDirectory() as tmp:
        exporter.export_csv(SAMPLE_ROWS, Path(tmp) / "out.csv", on_progress=calls.append)
    assert calls[-1] == 2
